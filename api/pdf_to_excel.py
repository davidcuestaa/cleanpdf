"""
API endpoint: POST /api/pdf_to_excel
Extrae tablas de PDF a XLSX.

Entorno: Vercel serverless (100% Python, sin binarios del sistema).

Comportamiento:
- TODAS las tablas de TODAS las páginas van a UNA SOLA hoja "Datos"
- Entre tablas: 1 fila vacía + fila de título "Página X – Tabla Y"
- Si no hay tablas con bordes, usa extracción por espacios (facturas, listados)
- Styling: cabecera azul, filas alternas, columnas auto-ajustadas, freeze row 1

Instalar: pip install pdfplumber openpyxl pdfminer.six
"""
import io
import os
import cgi
import json
from http.server import BaseHTTPRequestHandler

try:
    import pdfplumber
    PDFPLUMBER_OK = True
except ImportError:
    PDFPLUMBER_OK = False

try:
    from pdfminer.high_level import extract_text as pm_extract_text
    PDFMINER_OK = True
except ImportError:
    PDFMINER_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


def _parse_multipart(handler):
    ct = handler.headers.get("Content-Type", "")
    if "multipart/form-data" not in ct:
        return None, None, "Content-Type must be multipart/form-data"
    env = {"REQUEST_METHOD": "POST", "CONTENT_TYPE": ct,
           "CONTENT_LENGTH": handler.headers.get("Content-Length", "0")}
    body = handler.rfile.read(int(handler.headers.get("Content-Length", 0)))
    fs = cgi.FieldStorage(fp=io.BytesIO(body), environ=env, keep_blank_values=True)
    if "file" not in fs:
        return None, None, "No 'file' field in request"
    f = fs["file"]
    return f.file.read(), f.filename, None


# ── Estilos ───────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="4361EE")
ALT_FILL      = PatternFill("solid", fgColor="EEF0FB")
SECTION_FILL  = PatternFill("solid", fgColor="E8EAF6")  # gris-azulado para títulos de sección
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
SECTION_FONT  = Font(bold=True, color="1A237E", name="Calibri", size=9, italic=True)
BODY_FONT     = Font(name="Calibri", size=10)
CENTER_AL     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_AL       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN          = Side(style="thin", color="C5CAE9")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header_row(ws, row_idx, n_cols):
    """Aplica estilo de cabecera (azul) a una fila."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.border = BORDER
        cell.alignment = CENTER_AL


def _style_body_row(ws, row_idx, n_cols, alt=False):
    """Aplica estilo de cuerpo a una fila."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        if alt:
            cell.fill = ALT_FILL
        cell.font   = BODY_FONT
        cell.border = BORDER
        cell.alignment = LEFT_AL


def _style_section_row(ws, row_idx, n_cols, label):
    """Fila de título de sección entre tablas."""
    cell = ws.cell(row=row_idx, column=1, value=label)
    cell.fill      = SECTION_FILL
    cell.font      = SECTION_FONT
    cell.alignment = LEFT_AL
    # Merge across all columns
    if n_cols > 1:
        ws.merge_cells(
            start_row=row_idx, start_column=1,
            end_row=row_idx, end_column=n_cols
        )


def _auto_width(ws):
    """Ajusta ancho de columnas al contenido."""
    for col in ws.columns:
        letter  = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            try:
                if cell.value:
                    val = str(cell.value).split("\n")[0]
                    max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max(max_len + 2, 6), 55)


# ── Extracción de tablas ──────────────────────────────────────────────────────

# Configuraciones de extracción en orden de preferencia
_TABLE_SETTINGS = [
    # 1. Líneas explícitas estrictas (tablas con bordes bien definidos)
    {"vertical_strategy": "lines_strict",
     "horizontal_strategy": "lines_strict",
     "snap_tolerance": 4, "join_tolerance": 3},
    # 2. Líneas normales (bordes parciales)
    {"vertical_strategy": "lines",
     "horizontal_strategy": "lines",
     "snap_tolerance": 5, "join_tolerance": 4},
    # 3. Basado en texto/espacios (facturas, listados sin bordes)
    {"vertical_strategy": "text",
     "horizontal_strategy": "text",
     "snap_tolerance": 3,
     "min_words_vertical": 2,
     "min_words_horizontal": 1},
]


def _clean_table(table):
    """Limpia None, normaliza strings, elimina filas vacías."""
    cleaned = []
    for row in table:
        clean_row = [str(c).strip() if c is not None else "" for c in row]
        if any(c for c in clean_row):
            cleaned.append(clean_row)
    return cleaned


def _text_to_rows(lines):
    """Intenta estructurar líneas de texto en columnas por espacios múltiples."""
    import re
    has_multi = any(re.search(r"  +", l) for l in lines[:15])
    if has_multi:
        rows = [re.split(r"  +", l) for l in lines]
        rows = [[c.strip() for c in r if c.strip()] for r in rows]
        rows = [r for r in rows if r]
        if rows:
            max_c = max(len(r) for r in rows)
            rows  = [r + [""] * (max_c - len(r)) for r in rows]
            return rows
    return [[l] for l in lines if l.strip()]


def extract_all_tables(file_bytes):
    """
    Extrae todas las tablas del PDF.
    Devuelve lista de (page_num, table_idx, [[row_data]]).
    """
    results = []

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            page_tables = []

            # Intentar cada configuración hasta encontrar tablas
            for settings in _TABLE_SETTINGS:
                try:
                    tables = page.extract_tables(settings) or []
                    valid  = [_clean_table(t) for t in tables if t]
                    valid  = [t for t in valid if len(t) >= 1]
                    if valid:
                        page_tables = valid
                        break
                except Exception:
                    continue

            if page_tables:
                for t_idx, table in enumerate(page_tables, 1):
                    results.append((page_num, t_idx, table))
            else:
                # Sin tablas detectadas: intentar texto estructurado
                try:
                    text  = page.extract_text() or ""
                    lines = [l.strip() for l in text.split("\n") if l.strip()]
                    if lines:
                        rows = _text_to_rows(lines)
                        if rows:
                            results.append((page_num, 1, rows))
                except Exception:
                    pass

    return results


def extract_fallback_text(file_bytes):
    """Fallback pdfminer: texto plano en una sola columna."""
    text  = pm_extract_text(io.BytesIO(file_bytes)) or ""
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    if not lines:
        return []
    return [(1, 1, [[l] for l in lines])]


# ── Construcción del XLSX ─────────────────────────────────────────────────────




def pdf_to_xlsx(file_bytes):
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl no está instalado")

    if PDFPLUMBER_OK:
        tables = extract_all_tables(file_bytes)
    elif PDFMINER_OK:
        tables = extract_fallback_text(file_bytes)
    else:
        raise RuntimeError("Instala pdfplumber: pip install pdfplumber openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"

    if not tables:
        ws["A1"] = "No se encontraron tablas en este PDF."
        ws["A1"].font = Font(italic=True, color="888888", name="Calibri")
        ws.column_dimensions["A"].width = 45
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Concatenar todas las tablas en una sola tabla continua ───────────
    # Lógica:
    #   - La cabecera de la primera tabla → fila 1 con estilo azul
    #   - Si la siguiente tabla tiene la MISMA cabecera que la primera
    #     → se omite (es la misma tabla partida entre páginas)
    #   - Si la cabecera es distinta → se incluye como nueva cabecera
    #     con estilo azul (caso de tablas distintas en el mismo PDF)
    #   - Todas las filas de datos van seguidas sin huecos

    n_cols      = max(len(row) for _, _, tbl in tables for row in tbl)
    current_row = 1
    first_header = None   # cabecera de la primera tabla
    data_row_idx = 0      # contador global para filas alternas

    for entry_idx, (page_num, t_idx, table) in enumerate(tables):
        if not table:
            continue
        tbl_cols = max(len(row) for row in table)
        if tbl_cols == 0:
            continue

        for r_idx, row in enumerate(table):
            # Rellenar hasta n_cols para mantener alineación
            padded = row + [""] * (n_cols - len(row))

            is_header_row = r_idx == 0

            if is_header_row:
                if first_header is None:
                    # Primera cabecera → escribir siempre
                    first_header = padded
                elif padded == first_header:
                    # Cabecera repetida (misma tabla, siguiente página) → omitir
                    continue
                # Cabecera distinta → escribir como nueva cabecera
                for c_idx, val in enumerate(padded, 1):
                    ws.cell(row=current_row, column=c_idx, value=val)
                _style_header_row(ws, current_row, n_cols)
                current_row += 1
            else:
                # Fila de datos
                for c_idx, val in enumerate(padded, 1):
                    ws.cell(row=current_row, column=c_idx, value=val)
                _style_body_row(ws, current_row, n_cols, alt=(data_row_idx % 2 == 0))
                data_row_idx += 1
                current_row  += 1

    # ── Post-proceso ──────────────────────────────────────────────────────
    _auto_width(ws)
    ws.freeze_panes = "A2"

    # Filtro automático sobre toda la tabla
    end_col = get_column_letter(n_cols)
    end_row = current_row - 1
    ws.auto_filter.ref = f"A1:{end_col}{end_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── HTTP handler ──────────────────────────────────────────────────────────────

class handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self._cors()
        self.end_headers()

    def do_POST(self):
        try:
            file_bytes, filename, err = _parse_multipart(self)
            if err:
                self._error(400, err)
                return
            if not filename or not filename.lower().endswith(".pdf"):
                self._error(400, "Por favor sube un archivo PDF (.pdf)")
                return
            if not file_bytes:
                self._error(400, "El archivo está vacío")
                return

            xlsx_bytes = pdf_to_xlsx(file_bytes)
            out_name   = filename.rsplit(".", 1)[0] + ".xlsx"

            self.send_response(200)
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet",
            )
            self.send_header("Content-Disposition",
                             f'attachment; filename="{out_name}"')
            self.send_header("Content-Length", str(len(xlsx_bytes)))
            self._cors()
            self.end_headers()
            self.wfile.write(xlsx_bytes)

        except Exception as e:
            self._error(500, f"Error al convertir: {str(e).splitlines()[0]}")

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def _error(self, code, msg):
        body = json.dumps({"error": msg}).encode()
        self.send_response(code)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self._cors()
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, fmt, *args):
        pass
